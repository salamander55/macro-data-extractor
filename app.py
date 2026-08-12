"""
Macro Data Extractor — User Interface
--------------------------------------
A simple, professional-looking web interface (built with Streamlit) that
does exactly what macro_data_extractor.py does from the terminal:
lets the user pick indicator(s) and a date range, then generates an
Excel file in the same format as the source sheet.

Run it with:
    streamlit run app.py
"""

import io
import datetime

import openpyxl
import streamlit as st
import matplotlib.pyplot as plt

from macro_data_extractor import (
    INPUT_FILE,
    SHEET_NAME,
    load_indicators,
    load_date_columns,
    filter_date_columns,
    build_output,
)

st.set_page_config(page_title="Macro Data Extractor", page_icon="📊", layout="centered")

# Darken the dropdown option text — Streamlit's default is a light gray
# that can look faint/hazy, especially for the indicator list.
st.markdown(
    """
    <style>
    div[data-baseweb="popover"] li,
    div[data-baseweb="popover"] li * ,
    div[data-baseweb="select"] div {
        color: #1a1a1a !important;
        font-weight: 500 !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


st.title("📊 Macro Data Extractor")
st.caption(
    "Select the indicator(s) and date range you need. "
    "The output Excel file will match the same layout as the source sheet."
)


@st.cache_data
def load_data():
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    ws = wb[SHEET_NAME]
    indicators = load_indicators(ws)
    date_columns = load_date_columns(ws)
    return indicators, date_columns


try:
    indicators, date_columns = load_data()
except FileNotFoundError:
    st.error(
        f"Could not find '{INPUT_FILE}'. Make sure it's in the same folder as this app."
    )
    st.stop()

# Build a friendly label for each indicator, e.g. "Call Money Rate (%)"
indicator_labels = {}
for num, (row, name, unit) in indicators.items():
    label = f"{name} ({unit})" if unit else name
    indicator_labels[label] = num

# Build a friendly label for each month, e.g. "Jul 2009"
date_strings = [d.strftime("%b %Y") for _, d in date_columns]
date_lookup = {d.strftime("%b %Y"): d for _, d in date_columns}

st.subheader("1. Choose your indicator(s)")
selected_labels = st.multiselect(
    "Indicators",
    options=list(indicator_labels.keys()),
    placeholder="Start typing or click to choose...",
)

st.subheader("2. Choose your date range")
col1, col2 = st.columns(2)
with col1:
    start_label = st.selectbox("From", options=date_strings, index=0)
with col2:
    end_label = st.selectbox("To", options=date_strings, index=len(date_strings) - 1)

st.subheader("3. View charts (optional)")
st.caption("Shown here in the app only — charts are not saved into the Excel file.")
col3, col4 = st.columns(2)
with col3:
    show_bar = st.button("📊 Show Bar Chart", use_container_width=True)
with col4:
    show_line = st.button("📈 Show Line Chart", use_container_width=True)


def get_selection():
    """Read the current indicator + date selections and validate them.
    Returns (chosen_numbers, filtered_date_cols) or (None, None) if invalid."""
    if not selected_labels:
        st.error("Please select at least one indicator first.")
        return None, None

    start = date_lookup[start_label]
    end = date_lookup[end_label]
    if start > end:
        start, end = end, start

    chosen_numbers = [indicator_labels[label] for label in selected_labels]
    filtered_date_cols = filter_date_columns(date_columns, start, end)

    if not filtered_date_cols:
        st.error("No data found in that date range. Please pick a different range.")
        return None, None

    return chosen_numbers, filtered_date_cols


def get_series(indicator_row, filtered_date_cols, ws_in):
    """Return (month_labels, values) for one indicator, skipping months
    where the value is missing."""
    month_labels, values = [], []
    for col, d in filtered_date_cols:
        value = ws_in.cell(row=indicator_row, column=col).value
        if value is not None:
            month_labels.append(d.strftime("%b %y"))
            values.append(value)
    return month_labels, values


if show_bar or show_line:
    chosen_numbers, filtered_date_cols = get_selection()
    if chosen_numbers:
        wb_chart = openpyxl.load_workbook(INPUT_FILE, data_only=True)
        ws_chart = wb_chart[SHEET_NAME]

        for num in chosen_numbers:
            row, name, unit = indicators[num]
            months, values = get_series(row, filtered_date_cols, ws_chart)
            unit_label = unit if unit else "value"

            if not values:
                st.warning(f"No data available for **{name}** in this date range.")
                continue

            if show_bar:
                st.markdown(f"**{name}** — Bar Chart")
                fig, ax = plt.subplots(figsize=(8, 3.5))
                ax.bar(months, values, color="#4C78A8")
                ax.set_ylabel(unit_label)
                ax.tick_params(axis="x", rotation=45)
                fig.tight_layout()
                st.pyplot(fig)
                plt.close(fig)

            if show_line:
                st.markdown(f"**{name}** — Line Chart")
                fig, ax = plt.subplots(figsize=(8, 3.5))
                ax.plot(months, values, color="#4C78A8", marker="o", linewidth=2)
                ax.set_ylabel(unit_label)
                ax.tick_params(axis="x", rotation=45)
                ax.grid(axis="y", linestyle="--", alpha=0.4)
                fig.tight_layout()
                st.pyplot(fig)
                plt.close(fig)

st.subheader("4. Generate the file")
generate = st.button("Generate Excel", type="primary", use_container_width=True)

if generate:
    chosen_numbers, filtered_date_cols = get_selection()
    if chosen_numbers:
        wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
        ws_in = wb[SHEET_NAME]
        wb_out = build_output(ws_in, indicators, chosen_numbers, filtered_date_cols)

        buffer = io.BytesIO()
        wb_out.save(buffer)
        buffer.seek(0)

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Macro_Output_{timestamp}.xlsx"

        st.success(
            f"Done! {len(chosen_numbers)} indicator(s) x {len(filtered_date_cols)} months ready."
        )
        st.download_button(
            label="⬇️ Download Excel file",
            data=buffer,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )