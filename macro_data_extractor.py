"""
Macro Data Extractor
---------------------
Reads the "Monthly" sheet of the macroeconomic data file, asks the user
which indicator(s) and date range they want, and writes the result to a
new Excel file in the same row/column layout as the source sheet.
"""

import datetime
import openpyxl

# ----------------------------- SETTINGS ------------------------------
INPUT_FILE = "monthly_macro.xlsx"   # must be in the same folder as this script
SHEET_NAME = "Monthly"
DATE_ROW = 3          # row that holds the actual month dates
DATA_START_ROW = 4    # first row that holds indicator data
NAME_COL = 1           # column A
UNIT_COL = 2           # column B
SOURCE_COL = 3         # column C
FIRST_DATE_COL = 4     # column D onward holds monthly values
# -----------------------------------------------------------------------


def load_indicators(ws):
    """Return {display_number: (row_number, name, unit)} for every
    indicator row that has a name in column A."""
    indicators = {}
    counter = 1
    for row in range(DATA_START_ROW, ws.max_row + 1):
        name = ws.cell(row=row, column=NAME_COL).value
        if name is None or str(name).strip() == "":
            continue
        unit = ws.cell(row=row, column=UNIT_COL).value
        indicators[counter] = (row, name, unit)
        counter += 1
    return indicators


def load_date_columns(ws):
    """Return a list of (column_number, date) for every column that has
    a real date in the date row."""
    date_cols = []
    for col in range(FIRST_DATE_COL, ws.max_column + 1):
        value = ws.cell(row=DATE_ROW, column=col).value
        if isinstance(value, datetime.datetime):
            date_cols.append((col, value.date()))
    return date_cols


def show_indicator_menu(indicators):
    print("\nAvailable macro indicators:\n")
    for num, (row, name, unit) in indicators.items():
        unit_display = f" ({unit})" if unit else ""
        print(f"  {num:>2}. {name}{unit_display}")
    print()


def ask_indicator_choice(indicators):
    show_indicator_menu(indicators)
    while True:
        raw = input(
            "Enter the number(s) of the indicator(s) you want "
            "(comma-separated, e.g. 1,4,21): "
        ).strip()
        try:
            chosen_numbers = [int(x.strip()) for x in raw.split(",") if x.strip() != ""]
        except ValueError:
            print("Please enter numbers only, separated by commas. Try again.\n")
            continue

        invalid = [n for n in chosen_numbers if n not in indicators]
        if invalid:
            print(f"These numbers don't match the list: {invalid}. Try again.\n")
            continue

        if not chosen_numbers:
            print("Please enter at least one number.\n")
            continue

        return chosen_numbers


def ask_date(prompt_text):
    while True:
        raw = input(prompt_text).strip()
        try:
            # Accepts formats like "07-2009" or "2009-07"
            if "-" in raw and len(raw.split("-")[0]) == 4:
                return datetime.datetime.strptime(raw, "%Y-%m").date()
            else:
                return datetime.datetime.strptime(raw, "%m-%Y").date()
        except ValueError:
            print("Please use MM-YYYY format, e.g. 07-2015. Try again.\n")


def ask_date_range():
    print("\nEnter the date range you want (monthly data, format MM-YYYY).")
    start = ask_date("From (MM-YYYY): ")
    end = ask_date("To   (MM-YYYY): ")
    if start > end:
        start, end = end, start
    return start, end


def filter_date_columns(date_columns, start, end):
    return [(col, d) for col, d in date_columns if start <= d <= end]


def build_output(ws_in, indicators, chosen_numbers, filtered_date_cols):
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Monthly"

    # Header row: Name, Unit, Source, then each selected month
    ws_out.cell(row=1, column=NAME_COL, value="Name")
    ws_out.cell(row=1, column=UNIT_COL, value="Unit")
    ws_out.cell(row=1, column=SOURCE_COL, value="Source")
    for i, (src_col, d) in enumerate(filtered_date_cols):
        out_col = FIRST_DATE_COL + i
        cell = ws_out.cell(row=1, column=out_col, value=d)
        cell.number_format = "mmm-yy"

    # Data rows
    out_row = 2
    for num in chosen_numbers:
        src_row, name, unit = indicators[num]
        source_val = ws_in.cell(row=src_row, column=SOURCE_COL).value
        ws_out.cell(row=out_row, column=NAME_COL, value=name)
        ws_out.cell(row=out_row, column=UNIT_COL, value=unit)
        ws_out.cell(row=out_row, column=SOURCE_COL, value=source_val)
        for i, (src_col, d) in enumerate(filtered_date_cols):
            out_col = FIRST_DATE_COL + i
            value = ws_in.cell(row=src_row, column=src_col).value
            ws_out.cell(row=out_row, column=out_col, value=value)
        out_row += 1

    return wb_out


def main():
    print("Loading data from", INPUT_FILE, "...")
    wb_in = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    ws_in = wb_in[SHEET_NAME]

    indicators = load_indicators(ws_in)
    date_columns = load_date_columns(ws_in)

    chosen_numbers = ask_indicator_choice(indicators)
    start, end = ask_date_range()
    filtered_date_cols = filter_date_columns(date_columns, start, end)

    if not filtered_date_cols:
        print("No data found in that date range. Please check the dates and try again.")
        return

    wb_out = build_output(ws_in, indicators, chosen_numbers, filtered_date_cols)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_name = f"Macro_Output_{timestamp}.xlsx"
    wb_out.save(output_name)

    print(f"\nDone! Your data has been saved to: {output_name}")


if __name__ == "__main__":
    main()
